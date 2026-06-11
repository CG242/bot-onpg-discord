"""Export Excel des données FT Championship."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

import config
from database import Database
from ranking import format_tier, win_ratio


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 50)


def _header_row(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)


def build_season_excel(db: Database, season_id: int, path: Path) -> Path:
    season = db.get_season(season_id) or db.get_active_season()
    season_name = (season or {}).get("name", f"saison_{season_id}")

    wb = Workbook()

    # --- Classement ---
    ws_lb = wb.active
    ws_lb.title = "Classement"
    _header_row(
        ws_lb,
        ["Position", "Pseudo", "Rang", "Points", "Victoires", "Défaites", "Winrate %", "Région"],
    )
    rows = db.get_leaderboard(season_id, active_only=False)
    for pos, row in enumerate(rows, start=1):
        wins = int(row.get("ft_wins") or 0)
        losses = int(row.get("ft_losses") or 0)
        ws_lb.append([
            pos,
            row.get("display_name") or row.get("name"),
            format_tier(row.get("tier_rank")),
            int(row.get("elo") or 0),
            wins,
            losses,
            round(win_ratio(wins, losses), 1),
            (row.get("region") or "—").upper(),
        ])
    _auto_width(ws_lb)

    # --- Joueurs ---
    ws_p = wb.create_sheet("Joueurs")
    _header_row(
        ws_p,
        ["ID", "Pseudo", "Rang", "Points", "Région", "Discord ID", "Rang manuel"],
    )
    for p in db.list_all_players():
        ws_p.append([
            p["id"],
            db.player_display_name(p),
            format_tier(p.get("tier_rank")),
            int(p.get("elo") or 0),
            (p.get("region") or "—").upper(),
            p.get("discord_id") or "",
            "Oui" if p.get("rank_manual") else "Non",
        ])
    _auto_width(ws_p)

    # --- Matchs ---
    ws_m = wb.create_sheet("Matchs")
    _header_row(
        ws_m,
        [
            "Date", "Joueur 1", "Région J1", "Score J1", "Score J2",
            "Joueur 2", "Région J2", "FT", "Gagnant", "Inter-villes",
        ],
    )
    for m in db.list_season_matches_detailed(season_id):
        inter = (
            "Oui"
            if m.get("player1_region") in config.VALID_REGIONS
            and m.get("player2_region") in config.VALID_REGIONS
            and m.get("player1_region") != m.get("player2_region")
            else "Non"
        )
        ws_m.append([
            m.get("created_at"),
            m.get("player1_name"),
            (m.get("player1_region") or "—").upper(),
            m.get("score1"),
            m.get("score2"),
            m.get("player2_name"),
            (m.get("player2_region") or "—").upper(),
            f"FT{m.get('ft_type')}",
            m.get("winner_name"),
            inter,
        ])
    _auto_width(ws_m)

    # --- BZ vs PN ---
    ws_iv = wb.create_sheet("BZ vs PN")
    inter = db.get_inter_region_stats(season_id)
    _header_row(ws_iv, ["Indicateur", "Valeur"])
    ws_iv.append(["Saison", season_name])
    ws_iv.append(["Matchs inter-villes", inter["total_matches"]])
    ws_iv.append(["Victoires BZ", inter["bz_wins"]])
    ws_iv.append(["Victoires PN", inter["pn_wins"]])
    ws_iv.append([])
    _header_row(ws_iv, ["Joueur", "Région", "Victoires inter-villes", "Défaites inter-villes"])
    for row in inter.get("player_stats", []):
        ws_iv.append([
            row["name"],
            row["region"],
            row["wins"],
            row["losses"],
        ])
    _auto_width(ws_iv)

    # --- Saison ---
    ws_s = wb.create_sheet("Saison")
    _header_row(ws_s, ["Champ", "Valeur"])
    if season:
        ws_s.append(["Nom", season.get("name")])
        ws_s.append(["Début", season.get("start_date")])
        ws_s.append(["Fin", season.get("end_date") or "—"])
        ws_s.append(["Statut", season.get("status") or "active"])
    ws_s.append(["Exporté le", datetime.now().strftime("%d/%m/%Y %H:%M")])
    _auto_width(ws_s)

    path.parent.mkdir(parents=True, exist_ok=True)
    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in season_name)
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")
    wb.save(path)
    return path


def list_backup_files() -> list[Path]:
    backup_dir = Path(config.BACKUP_DIR)
    if not backup_dir.exists():
        return []
    files = sorted(
        backup_dir.glob("backup_*.json"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return files
